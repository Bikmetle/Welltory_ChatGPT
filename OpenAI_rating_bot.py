import openai
import pandas as pd
from openpyxl import load_workbook
from openai_api_key import api_key

# Set up OpenAI API credentials
openai.api_key = api_key

# Define function to rate paragraphs using OpenAI API
def rate_paragraph(paragraph):
    response = openai.Completion.create(
        engine="text-davinci-002",
        prompt=f"Rate the following paragraph on a scale of 1-10, where 10 is the most positive and 1 is the most negative:\n\n{paragraph}\n\nRating:",
        temperature=0.5,
        max_tokens=1,
        n=1,
        stop=None,
        timeout=15,
    )
    rating = int(response.choices[0].text.strip())
    return rating

# Load Excel file and select sheet
filename = 'example.xlsx'
wb = load_workbook(filename)
sheet = wb['Data']

# # Get paragraphs from sheet and rate them using OpenAI API
paragraphs = sheet['B2:B11']
ratings = [rate_paragraph(p[0].value) for p in paragraphs]

# # Add ratings to sheet
for i, rating in enumerate(ratings):
    sheet.cell(row=i+2, column=4).value = rating

# Save Excel file
wb.save(filename)

# # Sort and save analyzed data to CSV file
data = pd.read_excel(filename, sheet_name='Data')
sorted_data = data.sort_values(by='rate', ascending=False)
sorted_data.to_csv(f'{filename.split(".")[0]}_analyzed.csv', index=False)
